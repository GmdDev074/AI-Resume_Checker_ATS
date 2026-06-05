"""Usability study persistence, eligibility, and SUS scoring."""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import streamlit as st

from resume_analyzer.services.storage.database_service import DatabaseService

# Standard System Usability Scale (SUS) — 1 = Strongly disagree, 5 = Strongly agree
SUS_QUESTIONS: List[str] = [
    "I would like to use this system frequently.",
    "I found the system unnecessarily complex.",
    "I thought the system was easy to use.",
    "I think I would need technical support to use this system.",
    "I found the various functions of this system were well integrated.",
    "I thought there was too much inconsistency in this system.",
    "I would imagine that most people would learn to use this system quickly.",
    "I found the system very cumbersome to use.",
    "I felt very confident using the system.",
    "I needed to learn a lot before I could get going with this system.",
]

# Reverse-scored SUS items (even-numbered questions, 1-based index)
SUS_REVERSE_ITEMS = {2, 4, 6, 8, 10}

TASK_LABELS: List[str] = [
    "Upload resume (PDF or sample)",
    "Enter or select a job description",
    "Run analysis and view scores",
    "Review skill gaps and recommendations",
    "Generate and export a PDF report",
]

LIKERT_QUESTIONS: List[str] = [
    "The layout is clear and professional.",
    "I understood the ATS score and breakdown.",
    "Skill gap information was useful.",
    "Recommendations were actionable.",
    "The app responded in acceptable time.",
    "I would recommend this tool to a job seeker.",
]


@dataclass(frozen=True)
class UsabilitySummary:
    """Aggregated usability metrics for thesis reporting."""

    response_count: int
    task_success_rates: Dict[str, float]
    mean_likert: Dict[str, float]
    mean_sus: float
    mean_task_times_sec: Dict[str, Optional[float]]


def compute_sus_score(ratings: List[int]) -> float:
    """
    Compute SUS score (0–100) from ten 1–5 Likert responses.

    Args:
        ratings: Exactly ten integers from 1 to 5.

    Returns:
        SUS score between 0 and 100.
    """
    if len(ratings) != 10:
        raise ValueError("SUS requires exactly 10 ratings")
    total = 0.0
    for idx, rating in enumerate(ratings, start=1):
        if not 1 <= rating <= 5:
            raise ValueError(f"SUS rating must be 1–5, got {rating}")
        if idx in SUS_REVERSE_ITEMS:
            total += 5 - rating
        else:
            total += rating - 1
    return round(total * 2.5, 2)


def sync_report_unlock_from_session() -> None:
    """
    Mark workflow complete when Reports page generated a PDF in this session.

    Requires last_report_path (set only on Generate PDF) plus prior steps.
    """
    if st.session_state.get("report_generated"):
        return
    report_path = st.session_state.get("last_report_path")
    if not report_path or not Path(str(report_path)).is_file():
        return
    if not (
        st.session_state.get("resumes")
        and st.session_state.get("job")
        and st.session_state.get("last_analysis")
    ):
        return
    st.session_state["report_generated"] = True


def is_usability_study_unlocked() -> bool:
    """
    Return True only after the user generated a PDF report on the Reports page.
    """
    sync_report_unlock_from_session()
    return bool(st.session_state.get("report_generated"))


def mark_report_generated() -> None:
    """Record that the PDF report step was completed (call after successful export)."""
    st.session_state["report_generated"] = True


def openpyxl_available() -> bool:
    """Return True if openpyxl is installed (required for Excel export)."""
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


class UsabilityService:
    """Save and summarize usability study responses."""

    def __init__(self, db: Optional[DatabaseService] = None) -> None:
        self._db = db or DatabaseService()

    def save_response(self, payload: Dict[str, Any]) -> int:
        """Persist one usability survey submission."""
        return self._db.save_usability_response(payload)

    def list_responses(self, limit: int = 500) -> List[Dict[str, Any]]:
        """Fetch stored responses newest first."""
        return self._db.get_usability_responses(limit=limit)

    def summarize(self) -> UsabilitySummary:
        """Build aggregate metrics for thesis tables."""
        rows = self.list_responses()
        if not rows:
            return UsabilitySummary(
                response_count=0,
                task_success_rates={},
                mean_likert={},
                mean_sus=0.0,
                mean_task_times_sec={},
            )

        task_keys = [f"task_t{i}_success" for i in range(1, 6)]
        time_keys = [f"task_t{i}_time_sec" for i in range(1, 6)]
        likert_keys = [f"likert_q{i}" for i in range(1, 7)]

        task_success_rates: Dict[str, float] = {}
        for i, key in enumerate(task_keys, start=1):
            successes = sum(1 for r in rows if r.get(key))
            task_success_rates[f"T{i}"] = round(100.0 * successes / len(rows), 1)

        mean_likert: Dict[str, float] = {}
        for i, key in enumerate(likert_keys, start=1):
            values = [r[key] for r in rows if r.get(key) is not None]
            if values:
                mean_likert[f"Q{i}"] = round(sum(values) / len(values), 2)

        sus_values = [r["sus_score"] for r in rows if r.get("sus_score") is not None]
        mean_sus = round(sum(sus_values) / len(sus_values), 2) if sus_values else 0.0

        mean_task_times: Dict[str, Optional[float]] = {}
        for i, key in enumerate(time_keys, start=1):
            times = [r[key] for r in rows if r.get(key) is not None and r[key] > 0]
            mean_task_times[f"T{i}"] = round(sum(times) / len(times), 1) if times else None

        return UsabilitySummary(
            response_count=len(rows),
            task_success_rates=task_success_rates,
            mean_likert=mean_likert,
            mean_sus=mean_sus,
            mean_task_times_sec=mean_task_times,
        )

    def build_export_payload(self) -> Dict[str, Any]:
        """JSON-serializable summary plus raw rows for thesis export."""
        summary = self.summarize()
        return {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "response_count": summary.response_count,
            "mean_sus": summary.mean_sus,
            "task_success_rates_pct": summary.task_success_rates,
            "mean_likert": summary.mean_likert,
            "mean_task_times_sec": summary.mean_task_times_sec,
            "responses": self.list_responses(),
        }

    def export_json_text(self) -> str:
        """Full export payload as formatted JSON string for download."""
        return json.dumps(self.build_export_payload(), indent=2, ensure_ascii=False)

    def _csv_fieldnames(self) -> List[str]:
        """Column order for CSV and Excel raw export."""
        return [
            "id",
            "participant_id",
            "role",
            "task_t1_success",
            "task_t2_success",
            "task_t3_success",
            "task_t4_success",
            "task_t5_success",
            "task_t1_time_sec",
            "task_t2_time_sec",
            "task_t3_time_sec",
            "task_t4_time_sec",
            "task_t5_time_sec",
            "likert_q1",
            "likert_q2",
            "likert_q3",
            "likert_q4",
            "likert_q5",
            "likert_q6",
            "sus_q1",
            "sus_q2",
            "sus_q3",
            "sus_q4",
            "sus_q5",
            "sus_q6",
            "sus_q7",
            "sus_q8",
            "sus_q9",
            "sus_q10",
            "sus_score",
            "comments",
            "analysis_id",
            "created_at",
        ]

    def export_csv_text(self) -> str:
        """Flat CSV of all responses (one row per participant) for thesis tables."""
        rows = self.list_responses()
        fieldnames = self._csv_fieldnames()
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})
        return buffer.getvalue()

    def export_excel_bytes(self) -> bytes:
        """
        Build Excel workbook with raw data, summary metrics, and embedded charts.

        Sheets: Responses, Summary, Charts (task %, Likert means, SUS by participant).

        Raises:
            ImportError: If openpyxl is not installed.
        """
        if not openpyxl_available():
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Run: pip install openpyxl"
            )
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        rows = self.list_responses()
        summary = self.summarize()
        fieldnames = self._csv_fieldnames()

        wb = Workbook()

        # --- Sheet 1: Responses ---
        ws_data = wb.active
        ws_data.title = "Responses"
        ws_data.append(fieldnames)
        for cell in ws_data[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws_data.append([row.get(k, "") for k in fieldnames])
        for col in range(1, len(fieldnames) + 1):
            ws_data.column_dimensions[get_column_letter(col)].width = 14

        # --- Sheet 2: Summary ---
        ws_sum = wb.create_sheet("Summary")
        ws_sum["A1"] = "Usability study export"
        ws_sum["A1"].font = Font(bold=True, size=14)
        ws_sum["A3"] = "Exported at (UTC)"
        ws_sum["B3"] = datetime.now(timezone.utc).isoformat()
        ws_sum["A4"] = "Response count"
        ws_sum["B4"] = summary.response_count
        ws_sum["A5"] = "Mean SUS (0-100)"
        ws_sum["B5"] = summary.mean_sus

        ws_sum["A7"] = "Task"
        ws_sum["B7"] = "Success %"
        ws_sum["A7"].font = ws_sum["B7"].font = Font(bold=True)
        r = 8
        for task_id in sorted(summary.task_success_rates.keys()):
            ws_sum.cell(r, 1, task_id)
            ws_sum.cell(r, 2, summary.task_success_rates[task_id])
            r += 1

        likert_start = r + 2
        ws_sum.cell(likert_start, 1, "Likert question").font = Font(bold=True)
        ws_sum.cell(likert_start, 2, "Mean (1-5)").font = Font(bold=True)
        lr = likert_start + 1
        for q_id in sorted(summary.mean_likert.keys()):
            q_num = int(q_id.replace("Q", "")) - 1
            label = LIKERT_QUESTIONS[q_num][:48] if q_num < len(LIKERT_QUESTIONS) else q_id
            ws_sum.cell(lr, 1, f"{q_id}: {label}")
            ws_sum.cell(lr, 2, summary.mean_likert[q_id])
            lr += 1

        # --- Sheet 3: Charts ---
        ws_ch = wb.create_sheet("Charts")
        ws_ch["A1"] = "Thesis charts — open in Excel; graphs update with data"
        ws_ch["A1"].font = Font(bold=True)

        # Chart 1: Task success rates
        ws_ch["A3"] = "Task"
        ws_ch["B3"] = "Success %"
        ws_ch["A3"].font = ws_ch["B3"].font = Font(bold=True)
        task_row = 4
        for task_id in sorted(summary.task_success_rates.keys()):
            label = TASK_LABELS[int(task_id.replace("T", "")) - 1][:40]
            ws_ch.cell(task_row, 1, f"{task_id} {label}")
            ws_ch.cell(task_row, 2, summary.task_success_rates[task_id])
            task_row += 1

        if summary.task_success_rates:
            chart_tasks = BarChart()
            chart_tasks.type = "col"
            chart_tasks.title = "Task success rate (%)"
            chart_tasks.y_axis.title = "Percent"
            chart_tasks.x_axis.title = "Task"
            data_ref = Reference(ws_ch, min_col=2, min_row=3, max_row=task_row - 1)
            cats_ref = Reference(ws_ch, min_col=1, min_row=4, max_row=task_row - 1)
            chart_tasks.add_data(data_ref, titles_from_data=True)
            chart_tasks.set_categories(cats_ref)
            chart_tasks.height = 10
            chart_tasks.width = 18
            ws_ch.add_chart(chart_tasks, "D3")

        # Chart 2: Mean Likert
        likert_chart_row = task_row + 3
        ws_ch.cell(likert_chart_row, 1, "Question").font = Font(bold=True)
        ws_ch.cell(likert_chart_row, 2, "Mean").font = Font(bold=True)
        lr2 = likert_chart_row + 1
        for q_id in sorted(summary.mean_likert.keys()):
            ws_ch.cell(lr2, 1, q_id)
            ws_ch.cell(lr2, 2, summary.mean_likert[q_id])
            lr2 += 1

        if summary.mean_likert:
            chart_likert = BarChart()
            chart_likert.type = "col"
            chart_likert.title = "Mean Likert scores (1-5)"
            chart_likert.y_axis.title = "Mean score"
            data_ref = Reference(ws_ch, min_col=2, min_row=likert_chart_row, max_row=lr2 - 1)
            cats_ref = Reference(ws_ch, min_col=1, min_row=likert_chart_row + 1, max_row=lr2 - 1)
            chart_likert.add_data(data_ref, titles_from_data=True)
            chart_likert.set_categories(cats_ref)
            chart_likert.height = 10
            chart_likert.width = 18
            ws_ch.add_chart(chart_likert, f"D{likert_chart_row}")

        # Chart 3: SUS per participant
        sus_start = lr2 + 3
        ws_ch.cell(sus_start, 1, "Participant").font = Font(bold=True)
        ws_ch.cell(sus_start, 2, "SUS score").font = Font(bold=True)
        sr = sus_start + 1
        for row in reversed(rows):
            ws_ch.cell(sr, 1, row.get("participant_id"))
            ws_ch.cell(sr, 2, row.get("sus_score"))
            sr += 1

        if rows:
            chart_sus = BarChart()
            chart_sus.type = "col"
            chart_sus.title = "SUS score by participant"
            chart_sus.y_axis.title = "SUS (0-100)"
            data_ref = Reference(ws_ch, min_col=2, min_row=sus_start, max_row=sr - 1)
            cats_ref = Reference(ws_ch, min_col=1, min_row=sus_start + 1, max_row=sr - 1)
            chart_sus.add_data(data_ref, titles_from_data=True)
            chart_sus.set_categories(cats_ref)
            chart_sus.height = 10
            chart_sus.width = 18
            ws_ch.add_chart(chart_sus, f"D{sus_start}")

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def thesis_summary_rows(self) -> List[Dict[str, Any]]:
        """Compact per-participant rows for display in the app."""
        rows = self.list_responses()
        result = []
        for row in rows:
            tasks_ok = sum(1 for i in range(1, 6) if row.get(f"task_t{i}_success"))
            result.append(
                {
                    "Participant": row.get("participant_id"),
                    "Role": row.get("role"),
                    "Tasks OK": f"{tasks_ok}/5",
                    "SUS": row.get("sus_score"),
                    "Submitted": (row.get("created_at") or "")[:19].replace("T", " "),
                }
            )
        return result
